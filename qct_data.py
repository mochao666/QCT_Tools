# -*- coding: utf-8 -*-
"""
从 PDT 构建 QCT 内存数据（按 Sheet 的行列表），并支持读取/写入 QCT Excel、导出 Comments。
行结构：前 10 列为 QCT 内容，第 11 列 Event，第 12 列 Category，第 13 列 RTF Combine（可选，用于 Comments 过滤）。
"""

from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import (
    SHEET_SDTM,
    SHEET_ADAM_TFL,
    OUTPUT_TYPE_SDTM,
    QCT_HEADERS_SDTM,
    QCT_HEADERS_ADAM_TFL,
)


# ---------------------------------------------------------------------------
# 工具函数：单元格取值与 PDT 行 -> QCT 行转换
# ---------------------------------------------------------------------------

def _normalize_cell_value(val):
    """将单元格值转为字符串；空/NaN 为 ''，日期格式化为 YYYY-MM-DD。"""
    if pd.isna(val):
        return ""
    if isinstance(val, (pd.Timestamp, datetime)):
        return val.strftime("%Y-%m-%d")
    return val


def _row_to_qct_values(headers_config, row: pd.Series) -> list:
    """根据 (QCT列名, PDT列名) 配置和 PDT 的一行，生成 QCT 该行 10 列的值列表。"""
    values = []
    for qct_col, pdt_col in headers_config:
        if pdt_col is None:
            values.append("")
        else:
            values.append(_normalize_cell_value(row.get(pdt_col)))
    return values


# 行内列索引：前 10 列为 QCT，11=Event，12=Category，13=RTF Combine（导出 Comments 时仅保留为 'Y' 的行）
QCT_NUM_COLUMNS = 10
EVENT_COL_INDEX = 10
CATEGORY_COL_INDEX = 11
RTF_COMBINE_COL_INDEX = 12


# ---------------------------------------------------------------------------
# 从 PDT 构建内存行数据
# ---------------------------------------------------------------------------

def build_qct_rows_from_pdt(pdt_df: pd.DataFrame, event_value: str = ""):
    """
    从已清洗的 PDT DataFrame 生成两个 Sheet 的行数据。
    每行末尾带 Event（索引 10）、Category（索引 11）；event_value 用于新生成或增加行时标记该批行的 Event。
    返回 (sdtm_rows, adam_tfl_rows)，每个为 list of list，顺序与 QCT_HEADERS_* 一致 + Event + Category。
    """
    sdtm_rows = []
    adam_tfl_rows = []
    for _, row in pdt_df.iterrows():
        output_type = str(row.get("Output Type", "")).strip().upper()
        raw_cat = row.get("Category", "")
        if raw_cat is None or (isinstance(raw_cat, float) and pd.isna(raw_cat)):
            raw_cat = ""
        category = str(raw_cat).strip()
        # 将 "nan"、"None" 等视为空；仅统计并加入 Category 非空的行
        if not category or category.upper() in ("NAN", "NONE", "NAT"):
            continue
        rtf_combine = str(row.get("RTF Combine", "Y")).strip().upper()
        if output_type == OUTPUT_TYPE_SDTM:
            sdtm_rows.append(_row_to_qct_values(QCT_HEADERS_SDTM, row) + [event_value, category, rtf_combine])
        else:
            adam_tfl_rows.append(_row_to_qct_values(QCT_HEADERS_ADAM_TFL, row) + [event_value, category, rtf_combine])
    return sdtm_rows, adam_tfl_rows


# 可编辑列在 QCT 行中的索引（0-based），供 GUI 过滤与编辑
EDITABLE_COL_QC_DESC = 2   # QC results description
EDITABLE_COL_DATE_OF_QC = 3
EDITABLE_COL_FOLLOWUP_NOTES = 9


# ---------------------------------------------------------------------------
# 从 QCT Excel 读取行数据
# ---------------------------------------------------------------------------

def _openpyxl_cell_value(cell):
    """从 openpyxl 单元格取值，空为 ''，日期格式化为 Y-m-d。"""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, (datetime, pd.Timestamp)):
        return v.strftime("%Y-%m-%d")
    return "" if pd.isna(v) else str(v)


def read_qct_workbook(path: str):
    """
    从 QCT Excel 文件读取两个 Sheet 的数据，返回 (sdtm_rows, adam_tfl_rows)。
    每行为 9 列 QCT 数据 + Category（导入时默认 "Output"）。
    会过滤掉「QC results description (e.g. details of findings or passed)」为空的行。
    """
    wb = load_workbook(path, data_only=True)
    sdtm_rows = []
    adam_tfl_rows = []

    first_event = ""

    def sheet_to_rows_adam(sheet_name, out_list):
        """ADaM：A=Event, B-K=10 列 QCT"""
        nonlocal first_event
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        for row_cells in ws.iter_rows(min_row=2, max_col=11):
            values = [_openpyxl_cell_value(c) for c in row_cells]
            if len(values) < 11:
                values.extend([""] * (11 - len(values)))
            if not first_event and values[0]:
                first_event = str(values[0]).strip()
            qct_vals = values[1:11]
            if len(qct_vals) < 10:
                qct_vals.extend([""] * (10 - len(qct_vals)))
            qct_vals = qct_vals[:10]
            qc_desc = (qct_vals[EDITABLE_COL_QC_DESC] or "").strip()
            if not qc_desc:
                continue
            event_val = str(values[0]).strip() if values[0] else ""
            out_list.append(qct_vals + [event_val, "Output"])

    def sheet_to_rows_sdtm(sheet_name, out_list):
        """SDTM：A=Event, B-J=9 列 QCT（无 C 列），读入后补空列以与内部 10 列一致"""
        nonlocal first_event
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        for row_cells in ws.iter_rows(min_row=2, max_col=10):
            values = [_openpyxl_cell_value(c) for c in row_cells]
            if len(values) < 10:
                values.extend([""] * (10 - len(values)))
            if not first_event and values[0]:
                first_event = str(values[0]).strip()
            # 9 列 → 在索引 1 插入空列（QC checklist-index）得到 10 列
            qct_vals = values[1:2] + [""] + values[2:10]
            if len(qct_vals) < 10:
                qct_vals.extend([""] * (10 - len(qct_vals)))
            qct_vals = qct_vals[:10]
            qc_desc = (qct_vals[EDITABLE_COL_QC_DESC] or "").strip()
            if not qc_desc:
                continue
            event_val = str(values[0]).strip() if values[0] else ""
            out_list.append(qct_vals + [event_val, "Output"])

    sheet_to_rows_sdtm(SHEET_SDTM, sdtm_rows)
    sheet_to_rows_adam(SHEET_ADAM_TFL, adam_tfl_rows)
    return sdtm_rows, adam_tfl_rows, first_event


# ---------------------------------------------------------------------------
# 导出 Comments 工作簿
# ---------------------------------------------------------------------------
# 表头：Event + 5 列中英双行 + Developers、Validators（G/H 列隐藏）
COMMENTS_HEADERS_ZH = ["表格编号", "标题", "问题描述", "记录日期", "记录人"]
COMMENTS_HEADERS_EN = ["TFL No.", "TFL Title", "Issue Description", "Date Issued", "Issued by"]
COMMENTS_HEADERS_COMBINED = ["Event"] + [f"{zh}\n{en}" for zh, en in zip(COMMENTS_HEADERS_ZH, COMMENTS_HEADERS_EN)] + ["Developers", "Validators"]
DEVELOPERS_COL_INDEX = 5   # Person Responsible for Resolution
VALIDATORS_COL_INDEX = 4   # QC programmer Name


def write_comments_workbook(sdtm_rows: list, adam_tfl_rows: list, path: str, event_value: str = ""):
    """将当前 QCT 中 Category=Output 的审阅意见导出为单表 Excel。第一列 Event 按 event_value 自动填充（与导入 PDT 时选择的 Event 一致）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "审阅意见"
    # 单行表头：每格内为「中文\n英文」，灰色背景、自动换行
    ws.append(COMMENTS_HEADERS_COMBINED)
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    wrap_align = Alignment(wrap_text=True)
    for c in range(1, len(COMMENTS_HEADERS_COMBINED) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = grey_fill
        cell.alignment = wrap_align
    # 列宽：Event、表格编号、标题、问题描述、记录日期、记录人、Developers、Validators
    col_widths = (12, 18, 60, 60, 16, 14, 28, 18)
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # 隐藏 G/H 列（Developers、Validators），数据仍保留
    ws.column_dimensions[get_column_letter(7)].hidden = True
    ws.column_dimensions[get_column_letter(8)].hidden = True
    # 只保留 Category=Output 的记录；若行带 RTF Combine（来自 PDT），仅保留为 'Y' 的行；A=Event 使用 event_value 自动填充
    def emit_output_rows(rows):
        for row in rows:
            if len(row) <= CATEGORY_COL_INDEX:
                continue
            cat = str(row[CATEGORY_COL_INDEX]).strip().upper()
            if cat != "OUTPUT":
                continue
            if len(row) > RTF_COMBINE_COL_INDEX and str(row[RTF_COMBINE_COL_INDEX]).strip().upper() != "Y":
                continue
            developers = row[DEVELOPERS_COL_INDEX] if len(row) > DEVELOPERS_COL_INDEX else ""
            validators = row[VALIDATORS_COL_INDEX] if len(row) > VALIDATORS_COL_INDEX else ""
            ws.append([event_value or "", row[0], row[1], "", "", "", developers, validators])
    emit_output_rows(sdtm_rows)
    emit_output_rows(adam_tfl_rows)
    wb.save(path)


# ---------------------------------------------------------------------------
# QCT 工作簿样式与表头常量（写入 Excel 时使用）
# ---------------------------------------------------------------------------
QCT_COL_WIDTHS = (12, 15, 60, 45, 15, 15, 20, 15, 32, 15, 28)       # ADaM：Event + 10 列
QCT_COL_WIDTHS_SDTM = (12, 15, 45, 15, 15, 20, 15, 32, 15, 28)     # SDTM：Event + 9 列（无 C 列）
QCT_HEADER_ROW_SDTM = [
    "SDTM Datasets",
    "QC results description\n(e.g. details of findings or passed)",
    "Date of QC",
    "QC programmer\nName",
    "Person Responsible for Resolution if Findings",
    "Date of Resolved\nif Findings",
    "Resolution\nDetails",
    "Final Status if\nFindings",
    "Special Notes",
]
QCT_HEADER_ROW_ADAM = [
    "ADaM Dataset/\nTFL Number",
    "QC checklist-index",
    "QC results description\n(e.g. details of findings or passed)",
    "Date of QC",
    "QC programmer\nName",
    "Person Responsible for Resolution if Findings",
    "Date of Resolved\nif Findings",
    "Resolution\nDetails",
    "Final Status if\nFindings",
    "Special Notes",
]
FILL_LIGHT_BLUE = PatternFill(start_color="B8DCF9", end_color="B8DCF9", fill_type="solid")
FILL_LIGHT_GREEN = PatternFill(start_color="A8E0A8", end_color="A8E0A8", fill_type="solid")
HEADER_FONT = Font(bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BLACK_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
HEADER_BORDER = THIN_BLACK_BORDER

# List 表：Event 下拉选项、Final Status 选项、User List（从 PDT Users 表读）
SHEET_LIST_NAME = "List"
LIST_HEADERS = ["Event", "Final Status for\nFindings", "User List"]
LIST_COL_WIDTHS = (15, 22, 25)
EVENT_OPTIONS_LIST = [
    "CSR", "Dryrun", "IA", "DMC", "EOP2",
    "CSR1", "CSR2", "Dryrun1", "Dryrun2", "IA1", "IA2", "DMC1", "DMC2",
]
FINAL_STATUS_OPTIONS_LIST = ["Open", "Closed", "Follow up"]


def _read_users_from_pdt(pdt_path: str):
    """从 PDT 文件的 Users 表第一列读取内容，返回非空字符串列表。"""
    if not pdt_path or not isinstance(pdt_path, str):
        return []
    try:
        pdt_wb = load_workbook(pdt_path, data_only=True)
        ws = None
        for name in pdt_wb.sheetnames:
            if name.strip().upper() == "USERS":
                ws = pdt_wb[name]
                break
        if ws is None:
            return []
        users = []
        # 从第 2 行读起，跳过第一行表头（如 "User"），避免出现在 List 表第二行
        for row in ws.iter_rows(min_col=1, max_col=1, min_row=2):
            val = row[0].value
            if val is not None and str(val).strip():
                users.append(str(val).strip())
        return users
    except Exception:
        return []


# ---------------------------------------------------------------------------
# 写入 QCT 工作簿（两 Sheet + List + 数据验证）
# ---------------------------------------------------------------------------

def _set_qct_header_row(ws, headers: list):
    """设置 QCT 表头第一行：加粗、水平垂直居中、换行；仅 QC checklist-index 为浅绿，其余浅蓝；细黑边框；行高适应多行标题。"""
    ws.row_dimensions[1].height = 68
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c)
        cell.value = text
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER
        is_green = "QC checklist-index" in (text or "")
        cell.fill = FILL_LIGHT_GREEN if is_green else FILL_LIGHT_BLUE


def _set_data_row_borders(ws, num_cols: int):
    """从第 2 行起所有单元格设置黑色细边框。"""
    for r in range(2, ws.max_row + 1):
        for c in range(1, num_cols + 1):
            ws.cell(row=r, column=c).border = THIN_BLACK_BORDER


def write_qct_workbook(sdtm_rows: list, adam_tfl_rows: list, path: str, pdt_path: str = None, event_value: str = "", export_mode: str = "initial"):
    """将内存中的两个 Sheet 行数据写入 Excel。两 Sheet 第一列（A）为 Event，后 10 列为 QCT 内容；含 List 表。
    export_mode:
      - "initial"（初版QCT）: 所有行 A 列统一使用 event_value
      - "append"（新增Event）: 若行本身带 Event（row[EVENT_COL_INDEX] 非空）则使用该值；否则回退使用 event_value。"""
    use_single_event = export_mode == "initial"
    header_sdtm = ["Event"] + QCT_HEADER_ROW_SDTM
    header_adam = ["Event"] + QCT_HEADER_ROW_ADAM
    wb = Workbook()
    ws_sdtm = wb.active
    ws_sdtm.title = SHEET_SDTM
    _set_qct_header_row(ws_sdtm, header_sdtm)
    # SDTM 不写 C 列（QC checklist-index），即 row[1] 跳过
    for row in sdtm_rows:
        if use_single_event:
            ev = event_value or ""
        else:
            row_event = ""
            if len(row) > EVENT_COL_INDEX:
                row_event = (row[EVENT_COL_INDEX] or "").strip()
            ev = row_event or (event_value or "")
        sdtm_row_data = [ev] + row[0:1] + row[2:QCT_NUM_COLUMNS]
        ws_sdtm.append(sdtm_row_data)
    for i, w in enumerate(QCT_COL_WIDTHS_SDTM, start=1):
        ws_sdtm.column_dimensions[get_column_letter(i)].width = w
    _set_data_row_borders(ws_sdtm, num_cols=len(header_sdtm))

    ws_adam = wb.create_sheet(title=SHEET_ADAM_TFL)
    _set_qct_header_row(ws_adam, header_adam)
    for row in adam_tfl_rows:
        if use_single_event:
            ev = event_value or ""
        else:
            row_event = ""
            if len(row) > EVENT_COL_INDEX:
                row_event = (row[EVENT_COL_INDEX] or "").strip()
            ev = row_event or (event_value or "")
        ws_adam.append([ev] + row[:QCT_NUM_COLUMNS])
    for i, w in enumerate(QCT_COL_WIDTHS, start=1):
        ws_adam.column_dimensions[get_column_letter(i)].width = w
    _set_data_row_borders(ws_adam, num_cols=len(header_adam))

    # List 表：表头 + A 列 Event 选项 + B 列 Final Status + C 列 User List（来自 PDT Users）
    ws_list = wb.create_sheet(title=SHEET_LIST_NAME)
    ws_list.row_dimensions[1].height = 40
    for c, text in enumerate(LIST_HEADERS, start=1):
        cell = ws_list.cell(row=1, column=c)
        cell.value = text
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.fill = FILL_LIGHT_BLUE
        cell.border = THIN_BLACK_BORDER
    for i, w in enumerate(LIST_COL_WIDTHS, start=1):
        ws_list.column_dimensions[get_column_letter(i)].width = w
    # A 列（Event）展示选项：CSR, Dryrun, IA, DMC, EOP2
    for r, opt in enumerate(EVENT_OPTIONS_LIST, start=2):
        ws_list.cell(row=r, column=1, value=opt)
    for r, opt in enumerate(FINAL_STATUS_OPTIONS_LIST, start=2):
        ws_list.cell(row=r, column=2, value=opt)
    user_list = _read_users_from_pdt(pdt_path) if pdt_path else []
    for r, user in enumerate(user_list, start=2):
        ws_list.cell(row=r, column=3, value=user)

    # 两 Sheet 的 A 列（Event）设为下拉，选项引用 List 表 A2:A6
    event_ref = f"{SHEET_LIST_NAME}!$A$2:$A${1 + len(EVENT_OPTIONS_LIST)}"
    dv_event = DataValidation(type="list", formula1=event_ref, allow_blank=True)
    dv_event.add("A2:A1000")
    ws_sdtm.add_data_validation(dv_event)
    ws_adam.add_data_validation(dv_event)
    list_ref = f"{SHEET_LIST_NAME}!$B$2:$B$4"
    dv_final_sdtm = DataValidation(type="list", formula1=list_ref, allow_blank=True)
    dv_final_sdtm.add("I2:I1000")
    ws_sdtm.add_data_validation(dv_final_sdtm)
    dv_final_adam = DataValidation(type="list", formula1=list_ref, allow_blank=True)
    dv_final_adam.add("J2:J1000")
    ws_adam.add_data_validation(dv_final_adam)
    if user_list:
        last_row = 1 + len(user_list)
        user_ref = f"{SHEET_LIST_NAME}!$C$2:$C${last_row}"
        dv_user_sdtm = DataValidation(type="list", formula1=user_ref, allow_blank=True)
        dv_user_sdtm.add("E2:E1000")
        dv_user_sdtm.add("F2:F1000")
        ws_sdtm.add_data_validation(dv_user_sdtm)
        dv_user_adam = DataValidation(type="list", formula1=user_ref, allow_blank=True)
        dv_user_adam.add("F2:F1000")
        dv_user_adam.add("G2:G1000")
        ws_adam.add_data_validation(dv_user_adam)

    wb.save(path)
